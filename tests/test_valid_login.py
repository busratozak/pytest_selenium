import pytest
import openpyxl
import json
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as ec
from selenium.webdriver.support.wait import WebDriverWait
from constants import globalConstants 


class TestValidLogin():
    def setup_method(self,method):
        self.driver = webdriver.Chrome()
        self.driver.get(globalConstants.url)
        self.driver.maximize_window()
        self.vars = {}

    def teardown_method(self,method):
        self.driver.quit()

    def get_data():
        excelFile = openpyxl.load_workbook("data/userdata.xlsx")
        sheet = excelFile["Sayfa1"] 
        rows = sheet.max_row
        data = []
        for i in range(2,rows+1):
            username = sheet.cell(i,1).value 
            password = sheet.cell(i,2).value
            data.append((username,password))
    
        return data
    
    @pytest.mark.parametrize("username,password",get_data())
    def test_valid_login(self,username,password):
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"user-name")))
        usernameInput.send_keys(username)
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"password")))
        passwordInput.send_keys(password)
        loginButton = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"login-button")))
        loginButton.click()
        self.driver.get("https://www.saucedemo.com/inventory.html")
        products = self.driver.find_elements(By.CLASS_NAME,"inventory_item_price")
        assert len(products) == 6 


